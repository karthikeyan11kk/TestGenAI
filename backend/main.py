import json, os, io, base64, re
from datetime import datetime
from bson import ObjectId
from fastapi import FastAPI, HTTPException, Query, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel
from typing import List, Optional
import certifi
from groq import AsyncGroq

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

app = FastAPI(title="TestGen AI v7", version="7.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True,
                   allow_methods=["*"], allow_headers=["*"])

GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "").strip().strip('"').strip("'")
GROQ_MODEL   = os.environ.get("GROQ_MODEL",   "").strip().strip('"').strip("'")
MONGO_URI    = os.environ.get("MONGO_URI",     "").strip().strip('"').strip("'")

_mongo     = AsyncIOMotorClient(
    MONGO_URI if MONGO_URI else "mongodb://localhost:27017",
    tls=True if (MONGO_URI and "mongodb.net" in MONGO_URI) else False,
    tlsAllowInvalidCertificates=False,
    tlsCAFile=certifi.where() if (MONGO_URI and "mongodb.net" in MONGO_URI) else None,
    serverSelectionTimeoutMS=10000, connectTimeoutMS=10000,
)
_db        = _mongo["testgen"]
users_col  = _db["users"]
opt_col    = _db["options"]
cfg_col    = _db["configs"]
hist_col   = _db["history"]
matrix_col = _db["matrix_history"]

groq_client = AsyncGroq(api_key=GROQ_API_KEY) if GROQ_API_KEY else None

ACTURIS_BASE = (
    "You are a senior QA automation engineer for Acturis Core insurance platform.\n"
    "Key screens: IT(Broker) Instance, NI(UW) Instance, Rating Notes, Electronic file, "
    "DQI Tool, Renewal Quote, Accept Wizard, Risk Details.\n"
    "Migrations: R1=1C-to-1C, M2=NIG-to-1C (NG prefix policies), M1=RSA-to-1C."
)

PRODUCT_RULES = {
    "Property Owner":{"R1":["Login IT(Broker) and NI(UW)","Create backdated Intact PO NB Policy","Process Renewal in NI(UW)","Verify RES006C/RES006D in Rating Notes","Verify flood score Read Only","Verify DQI XML Perils and Proximity","Accept Wizard - Accepted status","Verify Exclude from Auto Renewals","Test End"],"M2":["Login IT(Broker)","Create backdated NIG PO NB Policy - verify NG prefix","Verify Autorenewal triggered","Verify original NGnnn number retained","Verify Migrated RQ in new Intact Version","Verify no claims","Verify Exclude from Auto Renewals changeable","Test End"],"M1":["Login IT(Broker)","Create backdated RSA PO Policy in 1C","Verify RSA reference mapped","Verify details match RSA source","Reconcile premiums","Verify no claims","Test End"]},
    "Shop":{"R1":["Login IT(Broker) and NI(UW)","Create backdated Intact Shops NB Policy","Process Renewal","Verify all standard covers","Verify endorsements in Rating Notes","Verify flood score Read Only","Verify DQI XML","Accept wizard - Accepted status","Test End"],"M2":["Login IT(Broker)","Create backdated NIG Shops NB Policy - verify NG prefix","Verify Autorenewal triggered","Verify TRA104/TRA111/TRA112 endorsements","Verify original NG policy number retained","Verify Migrated to new Intact Version","Verify no claims","Test End"]},
    "Office":{"R1":["Login IT(Broker) and NI(UW)","Create backdated Intact Office NB Policy","Process Renewal","Verify covers and endorsements","Verify flood score Read Only","Accept wizard - Accepted status","Test End"],"M1":["Login IT(Broker)","Create backdated RSA Office Policy in 1C","Verify RSA reference mapped","Verify office details match RSA","Reconcile premiums","Verify no claims","Test End"]},
    "TP":{"R1":["Login IT(Broker) and NI(UW)","Create backdated Intact TP NB Policy","Process Renewal","Verify all covers and endorsements","Accept wizard - Accepted status","Test End"],"M2":["Login IT(Broker)","Create backdated NIG TP NB Policy - verify NG prefix","Verify Autorenewal triggered","Verify original NG number retained","Verify Migrated to new Intact Version","Verify no claims","Test End"]},
}

@app.on_event("startup")
async def seed():
    try:
        for u in [{"email":"admin@testgen.com","role":"admin","name":"Admin"},
                  {"email":"user@testgen.com","role":"user","name":"Test User"}]:
            await users_col.update_one({"email":u["email"]},
                {"$setOnInsert":{**u,"created_at":datetime.utcnow().isoformat()}},upsert=True)
        if not await opt_col.find_one({}):
            await opt_col.insert_one({"products":["Property Owner","Shop","Office","TP"],
                "migrations":[{"id":"R1","label":"R1 - 1C TO 1C"},{"id":"M2","label":"M2 - NIG TO 1C"},{"id":"M1","label":"M1 - RSA TO 1C"}]})
        if not await cfg_col.find_one({}):
            docs=[]
            for prod,migs in PRODUCT_RULES.items():
                for mig,rules in migs.items():
                    docs.append({"key":f"{prod}__{mig}","product":prod,"migration":mig,"system_context":ACTURIS_BASE,"rules":rules})
            if docs: await cfg_col.insert_many(docs)
    except Exception as e:
        print(f"Seed warning: {e}")

class LoginReq(BaseModel):    email: str
class UserCreate(BaseModel):  email: str; name: str; role: str="user"
class MigModel(BaseModel):    id: str; label: str
class Options(BaseModel):     products: List[str]; migrations: List[MigModel]
class Config(BaseModel):      product:str; migration:str; system_context:str=""; rules:List[str]=[]
class GenReq(BaseModel):      product:str; migration:str; acceptance_criteria:List[str]; user_email:str; history_id:str=""
class RefineReq(BaseModel):   product:str; migration:str; user_email:str; feedback:str; previous_steps:list; previous_title:str=""; previous_preconditions:str=""; acceptance_criteria:List[str]=[]; history_id:str=""

def fix_id(d):
    if d and "_id" in d: d["id"]=str(d.pop("_id"))
    return d
def check_llm():
    if groq_client is None: raise HTTPException(503,"Groq not configured. Set GROQ_API_KEY.")

async def call_llm(sys_p:str, usr_p:str, max_tokens:int=4096) -> str:
    check_llm()
    try:
        resp=await groq_client.chat.completions.create(model=GROQ_MODEL,
            messages=[{"role":"system","content":sys_p},{"role":"user","content":usr_p}],
            temperature=0.1,max_tokens=max_tokens)
        return resp.choices[0].message.content or ""
    except Exception as e:
        err=str(e).lower()
        original=str(e)
        if "401" in err or "invalid api key" in err or "authentication" in err or "expired" in err:
            raise HTTPException(401,
                "Groq API key expired or invalid. Get a new key at https://console.groq.com → API Keys, "
                "then update GROQ_API_KEY in Render environment variables and redeploy.")
        if "429" in err or "rate limit" in err or "quota" in err or "too many" in err:
            # Extract "Please try again in Xm Ys" from the error message
            import re as _re
            retry_match = _re.search(r'[Pp]lease try again in (\d+(?:m\d+(?:\.\d+)?s?|\.\d+s|s))', original)
            retry_info = f" Please try again in {retry_match.group(1).strip()}." if retry_match else " Please wait a minute and try again."
            raise HTTPException(429, f"Groq rate limit reached.Try Again Tomorrow")
        raise HTTPException(502,f"Groq error: {e}")

def repair_json(raw:str) -> str:
    text=raw.strip()
    for fence in ["```json","```JSON","```"]:
        if fence in text:
            parts=text.split(fence); text=parts[1] if len(parts)>1 else text; break
    text=text.strip()
    for p in ["here is","here's","output:","result:","json:","answer:","sure!","certainly!"]:
        if text.lower().startswith(p):
            nl=text.find("\n")
            if nl>0: text=text[nl:].strip()
    # single-quote dicts
    if text.startswith("{'") or text.startswith("['"):
        try:
            import ast as _a; obj=_a.literal_eval(text); return json.dumps(obj)
        except: pass
    text="".join(c for c in text if c>=" " or c in "\n\r\t")
    start=next((i for i,c in enumerate(text) if c in "{["),-1)
    if start==-1: raise ValueError(f"No JSON. Model said: {text[:150]!r}")
    text=text[start:]
    depth=0; in_s=False; esc=False; end=-1
    op=text[0]; cl="}" if op=="{" else "]"
    for i,ch in enumerate(text):
        if esc: esc=False; continue
        if ch=="\\" and in_s: esc=True; continue
        if ch=='"' and not esc: in_s=not in_s; continue
        if in_s: continue
        if ch==op: depth+=1
        elif ch==cl:
            depth-=1
            if depth==0: end=i+1; break
    if end>0: text=text[:end]
    text=re.sub(r",\s*([}\]])",r"\1",text)
    try: json.loads(text); return text
    except:
        ob=text.count("{")-text.count("}"); ob2=text.count("[")-text.count("]")
        if text.count('"')%2!=0: text+='"'
        text+="]"*max(ob2,0)+"}"*max(ob,0)
        try: json.loads(text); return text
        except: return text+"}]}"

async def llm_json(sys_p:str, usr_p:str, max_tokens:int=4096) -> dict:
    def force_parse(raw):
        if not raw or not raw.strip(): return {}
        try: return json.loads(repair_json(raw))
        except: pass
        try:
            import ast as _a
            m=re.search(r"\{.*\}",raw,re.DOTALL)
            if m: return json.loads(json.dumps(_a.literal_eval(m.group())))
        except: pass
        try:
            f=re.sub(r"'([^']+)'(\s*:)",r'"\1"\2',raw)
            f=re.sub(r":\s*'([^']*)'",r': "\1"',f)
            f=re.sub(r",\s*([}\]])",r"\1",f)
            return json.loads(repair_json(f))
        except: pass
        return {}

    raw1=await call_llm(sys_p,usr_p,max_tokens)
    r1=force_parse(raw1)
    if r1 and (r1.get("steps") or r1.get("scenarios") or r1.get("mappings")): return r1

    raw2=await call_llm(
        "Output ONLY valid JSON. No text. No markdown. No single quotes. Start with {.",
        f'Fill this JSON with real content:\n{{"testTitle":"<t>","preconditions":"<p>","steps":[{{"stepNo":1,"step":"<s>","expectedResult":"<e>"}}]}}\nContext: {usr_p[:500]}')
    r2=force_parse(raw2)
    if r2 and (r2.get("steps") or r2.get("scenarios")): return r2

    raw3=await call_llm("Convert to valid JSON with double quotes. Output ONLY JSON.",f"Fix:\n{raw1[:800]}")
    r3=force_parse(raw3)
    if r3: return r3

    raise HTTPException(500,"Model could not produce valid JSON after 3 attempts. Try GROQ_MODEL=llama-3.1-8b-instant")

# Auth
@app.post("/api/auth/login")
async def login(req:LoginReq):
    user=await users_col.find_one({"email":req.email.strip().lower()})
    if not user: raise HTTPException(404,"Email not found.")
    return {"email":user["email"],"name":user.get("name",""),"role":user["role"]}

@app.get("/api/admin/users")
async def list_users():
    out=[]
    async for u in users_col.find({}).sort("created_at",-1): u.pop("_id",None); out.append(u)
    return out

@app.post("/api/admin/users")
async def add_user(user:UserCreate):
    email=user.email.strip().lower()
    if await users_col.find_one({"email":email}): raise HTTPException(400,"User exists")
    await users_col.insert_one({"email":email,"name":user.name,"role":user.role,"created_at":datetime.utcnow().isoformat()})
    return {"status":"created"}

@app.delete("/api/admin/users/{email}")
async def del_user(email:str):
    r=await users_col.delete_one({"email":email.lower()})
    if r.deleted_count==0: raise HTTPException(404,"Not found")
    return {"status":"deleted"}

@app.get("/api/options")
async def get_options():
    doc=await opt_col.find_one({})
    if not doc: return {"products":[],"migrations":[]}
    doc.pop("_id",None); return doc

@app.post("/api/admin/options")
async def save_options(opts:Options):
    await opt_col.delete_many({}); await opt_col.insert_one(opts.dict()); return {"status":"saved"}

@app.get("/api/admin/config/{product}/{migration}")
async def get_config(product:str,migration:str):
    doc=await cfg_col.find_one({"key":f"{product}__{migration}"})
    if not doc: return {"product":product,"migration":migration,"system_context":"","rules":[]}
    doc.pop("_id",None); doc.pop("key",None); return doc

@app.post("/api/admin/config")
async def save_config(cfg:Config):
    key=f"{cfg.product}__{cfg.migration}"
    await cfg_col.update_one({"key":key},{"$set":{**cfg.dict(),"key":key,"updated_at":datetime.utcnow().isoformat()}},upsert=True)
    return {"status":"saved"}

@app.delete("/api/admin/config/{product}/{migration}")
async def del_config(product:str,migration:str):
    await cfg_col.delete_one({"key":f"{product}__{migration}"}); return {"status":"deleted"}

@app.get("/api/history")
async def get_history(user_email:str=Query(...),q:Optional[str]=Query(None),product:Optional[str]=Query(None),migration:Optional[str]=Query(None),limit:int=Query(50,le=200)):
    filt={"user_email":user_email.lower()}
    if product: filt["product"]=product
    if migration: filt["migration"]=migration
    if q: filt["$or"]=[{"testTitle":{"$regex":q,"$options":"i"}},{"acs_text":{"$regex":q,"$options":"i"}}]
    out=[]
    async for doc in hist_col.find(filt,{"steps":0}).sort("created_at",-1).limit(limit):
        doc["id"]=str(doc.pop("_id")); out.append(doc)
    return out

@app.get("/api/history/{doc_id}")
async def get_history_item(doc_id:str):
    try: oid=ObjectId(doc_id)
    except: raise HTTPException(400,"Invalid ID")
    doc=await hist_col.find_one({"_id":oid})
    if not doc: raise HTTPException(404,"Not found")
    return fix_id(doc)

@app.delete("/api/history/{doc_id}")
async def del_history(doc_id:str):
    try: oid=ObjectId(doc_id)
    except: raise HTTPException(400,"Invalid ID")
    await hist_col.delete_one({"_id":oid}); return {"status":"deleted"}

async def save_tc(result:dict,req_data:dict,history_id:str) -> str:
    doc={**result,"product":req_data["product"],"migration":req_data["migration"],
         "acs":req_data["acs"],"acs_text":" ".join(req_data["acs"]),
         "ac_count":len(req_data["acs"]),"step_count":len(result.get("steps",[])),
         "user_email":req_data["user_email"].lower(),"updated_at":datetime.utcnow().isoformat()}
    if req_data.get("extra"): doc.update(req_data["extra"])
    if history_id:
        try:
            oid=ObjectId(history_id)
            orig=await hist_col.find_one({"_id":oid},{"created_at":1})
            doc["created_at"]=(orig or {}).get("created_at",datetime.utcnow().isoformat())
            await hist_col.update_one({"_id":oid},{"$set":doc}); return history_id
        except: pass
    doc["created_at"]=datetime.utcnow().isoformat()
    ins=await hist_col.insert_one(doc); return str(ins.inserted_id)

@app.post("/api/generate")
async def generate(req:GenReq):
    check_llm()
    if not req.acceptance_criteria: raise HTTPException(400,"At least one AC required")
    cfg=(await cfg_col.find_one({"key":f"{req.product}__{req.migration}"})) or {}
    sys_ctx=cfg.get("system_context",ACTURIS_BASE); rules=cfg.get("rules",[])
    numbered_acs="\n".join(f"AC{i+1}: {ac}" for i,ac in enumerate(req.acceptance_criteria))
    sys_p=(
        f"You are a senior Acturis QA engineer.\n\n"
        f"ACTURIS DOMAIN KNOWLEDGE (reference only — do NOT copy as steps):\n{sys_ctx[:500]}\n"
        +(f"\nWORKFLOW REFERENCE:\n"+"\n".join(f"• {r}" for r in rules[:10])+"\n" if rules else "")
        +"YOUR TASK:\n"
        "Combine ALL acceptance criteria into ONE single end-to-end test case.\n"
        "Flow sequentially through every scenario. Do NOT create separate test cases.\n"
        "Use Acturis domain knowledge to write realistic steps. Steps must MATCH the ACs.\n\n"
        "RULES:\n"
        "1. ALL ACs covered in one continuous E2E test case.\n"
        "2. Use Acturis screens: IT(Broker) Instance, NI(UW) Instance, Rating Notes, DQI Tool.\n"
        "3. Every step: specific 'step' action AND specific 'expectedResult'. No empty fields.\n"
        "4. Start with login. End with Test End.\n"
        "5. Output ONLY raw JSON. Start with {.\n\n"
        'FORMAT: {"testTitle":"<E2E title>","preconditions":"<preconditions>","steps":['
        '{"stepNo":1,"step":"<action>","expectedResult":"<outcome>"}]}'
    )
    usr_p=(f"Product: {req.product} | Migration: {req.migration}\n\n"
           f"ACCEPTANCE CRITERIA — combine ALL into ONE E2E TC:\n{numbered_acs}\n\n"
           "Generate ONE end-to-end test case covering all scenarios. Start with {")
    try: result=await llm_json(sys_p,usr_p)
    except Exception as e: raise HTTPException(500,f"LLM error: {e}")
    for s in result.get("steps",[]):
        if not s.get("expectedResult") and s.get("step","").strip().lower()!="test end":
            s["expectedResult"]=s.pop("expected_result",None) or s.pop("expected",None) or "Step completed successfully"
    hid=await save_tc(result,{"product":req.product,"migration":req.migration,"acs":req.acceptance_criteria,"user_email":req.user_email},req.history_id)
    result["history_id"]=hid; return result

@app.post("/api/refine")
async def refine(req:RefineReq):
    check_llm()
    if not req.feedback.strip(): raise HTTPException(400,"Feedback required")
    cfg=(await cfg_col.find_one({"key":f"{req.product}__{req.migration}"})) or {}
    sys_ctx=cfg.get("system_context",ACTURIS_BASE)
    existing="\n".join(f"Step {s.get('stepNo','')}: {s.get('step','')} → {s.get('expectedResult','')}" for s in (req.previous_steps or []))
    acs_text="\n".join(f"AC{i+1}: {ac}" for i,ac in enumerate(req.acceptance_criteria)) if req.acceptance_criteria else "See previous steps"
    sys_p=(f"You are a senior Acturis QA engineer.\n\nDOMAIN KNOWLEDGE (reference only):\n{sys_ctx[:400]}\n\n"
           "Refine the existing test case based on user feedback. Keep good steps. Fix/add/remove per feedback.\n"
           "Every step needs 'step' and 'expectedResult'. Last step: Test End. Output ONLY JSON. Start with {.\n"
           'FORMAT: {"testTitle":"<t>","preconditions":"<p>","steps":[{"stepNo":1,"step":"<s>","expectedResult":"<e>"}]}')
    usr_p=(f"Product:{req.product} Migration:{req.migration}\nACs:{acs_text}\n\nCurrent steps:\n{existing}\n\nFeedback: {req.feedback}\n\nGenerate improved test case. Start with {{")
    try: result=await llm_json(sys_p,usr_p)
    except Exception as e: raise HTTPException(500,f"LLM error: {e}")
    for s in result.get("steps",[]):
        if not s.get("expectedResult") and s.get("step","").strip().lower()!="test end":
            s["expectedResult"]=s.pop("expected_result",None) or "Step completed successfully"
    hid=await save_tc(result,{"product":req.product,"migration":req.migration,"acs":req.acceptance_criteria,"user_email":req.user_email,"extra":{"refined":True,"feedback":req.feedback}},req.history_id)
    result["history_id"]=hid; return result

# Matrix History
@app.get("/api/matrix/history")
async def get_matrix_history(user_email:str=Query(...),limit:int=Query(30,le=100)):
    out=[]
    async for doc in matrix_col.find({"user_email":user_email.lower()},{"excel_bytes":0}).sort("created_at",-1).limit(limit):
        doc["id"]=str(doc.pop("_id")); out.append(doc)
    return out

@app.get("/api/matrix/history/{doc_id}/download")
async def download_matrix(doc_id:str):
    try: oid=ObjectId(doc_id)
    except: raise HTTPException(400,"Invalid ID")
    doc=await matrix_col.find_one({"_id":oid})
    if not doc: raise HTTPException(404,"Not found")
    b64=doc.get("excel_bytes","")
    if not b64: raise HTTPException(404,"File not stored")
    return Response(content=base64.b64decode(b64),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":f"attachment; filename={doc.get('filename','matrix.xlsx')}"})

@app.delete("/api/matrix/history/{doc_id}")
async def del_matrix(doc_id:str):
    try: oid=ObjectId(doc_id)
    except: raise HTTPException(400,"Invalid ID")
    await matrix_col.delete_one({"_id":oid}); return {"status":"deleted"}


@app.post("/api/matrix")
async def generate_matrix(file: UploadFile = File(...), user_email: str = Form(...)):
    check_llm()
    if not OPENPYXL_OK:
        raise HTTPException(500, "openpyxl not installed")

    content_bytes = await file.read()
    try:
        wb_in = openpyxl.load_workbook(io.BytesIO(content_bytes))
    except Exception as e:
        raise HTTPException(400, f"Cannot read Excel: {e}")
    ws_in = wb_in.active

    # ── Detect header row ─────────────────────────────────────────────────────
    data_start = 2
    for r in range(1, min(15, ws_in.max_row + 1)):
        vals = [str(ws_in.cell(r, c).value or "").lower() for c in range(1, 9)]
        if any(k in " ".join(vals) for k in ["functional", "req. id", "req id", "requirement"]):
            data_start = r + 1
            break

    # ── Read requirements ─────────────────────────────────────────────────────
    requirements = []
    for row in range(data_start, ws_in.max_row + 1):
        cols = [ws_in.cell(row, c).value for c in range(1, 9)]
        if not any(cols):
            continue
        req_id   = str(cols[3] or "")
        req_desc = str(cols[4] or "")
        req_stmt = str(cols[5] or "")
        acs      = str(cols[6] or "")
        if not req_id and not req_desc and not acs:
            continue
        requirements.append({
            "row": row, "col_a": str(cols[0] or ""), "col_b": str(cols[1] or ""),
            "col_c": str(cols[2] or ""), "req_id": req_id or f"ROW{row}",
            "req_desc": req_desc[:300], "req_stmt": req_stmt[:300],
            "acs": acs, "col_h": str(cols[7] or ""),
        })
    if not requirements:
        raise HTTPException(400, "No requirements found. Check Excel has Req ID in column D.")

    # ── Detect migration pattern ──────────────────────────────────────────────
    all_text = " ".join(r["col_c"] + " " + r["req_desc"] + " " + r["acs"][:100]
                        for r in requirements).lower()
    if "nig" in all_text or " ng" in all_text or "m2" in all_text:
        pattern = "M2"
    elif "rsa" in all_text or "m1" in all_text:
        pattern = "M1"
    else:
        pattern = "R1"

    # ── LLM helper ────────────────────────────────────────────────────────────
    def try_parse(raw):
        if not raw: return {}
        try: return json.loads(repair_json(raw))
        except: pass
        try:
            import ast as _a
            m = re.search(r"\{.*\}", raw, re.DOTALL)
            if m: return json.loads(json.dumps(_a.literal_eval(m.group())))
        except: pass
        return {}

    async def llm_call(sp, up, mt=2500):
        raw = await call_llm(sp, up, mt)
        r = try_parse(raw)
        if r: return r
        # retry once
        raw2 = await call_llm(
            "Output ONLY valid JSON. No explanation. Start with {",
            f"Return JSON for:\n{up[:400]}"
        )
        return try_parse(raw2) or {}

    # ═══════════════════════════════════════════════════════════════════════════
    # STEP 1 — Generate scenario headings
    # ═══════════════════════════════════════════════════════════════════════════
    # Build context from actual AC content
    ac_samples = "\n\n".join(
        f"[{r['req_id']}] {r['req_desc'][:60]}\nAC text: {r['acs'][:200]}"
        for r in requirements[:12]
    )

    sc_resp = await llm_call(
        "You are a QA analyst for Acturis insurance. Output ONLY JSON. Start with {.",
        (
            f"Analyse these Acturis {pattern} migration requirements and generate "
            f"8-21 distinct E2E test scenario headings.\n\n"
            "Rules for scenario names:\n"
            "- Each name describes ONE specific end-to-end Acturis test flow\n"
            "- Format: '[Renewal type] - [Specific flow description]'\n"
            "- Be specific: include key actions like 'MTA', 'Claims', 'Endorsements', 'Accept Wizard'\n"
            "- Example good names:\n"
            "  * 'Intact Renewal - Non-Risk Updates - Validate and Accept Renewal'\n"
            "  * 'Intact Renewal - Risk Update at Quote Level - ProductWriter Refers - UW Authorize'\n"
            "  * 'Intact Renewal - Perform MTA in Renewal Cycle and Reject'\n"
            "  * 'Intact Renewal - Add Pre Inception Claims on NB - Process Renewal'\n"
            "  * 'Intact Renewal - Verify Exclude from Auto Renewals Option'\n"
            "- Read the AC texts carefully to identify all distinct flows\n"
            "- Section = 'Core Functionality' for business flows, 'Tech Readiness' for data/XML\n\n"
            f"Requirements to analyse:\n{ac_samples}\n\n"
            "Return: {\"scenarios\":[{\"number\":1,\"name\":\"Intact Renewal - ...\","
            "\"section\":\"Core Functionality\"}]}\n"
            "Start with {"
        ),
        3000
    )
    scenarios = sc_resp.get("scenarios", [])
    if not scenarios:
        raise HTTPException(500, "Could not generate scenarios. Retry or check GROQ_API_KEY.")

    sc_list = "\n".join(f"{s['number']}. {s['name']}" for s in scenarios)

    # ═══════════════════════════════════════════════════════════════════════════
    # STEP 2 — Map requirements to scenarios in batches of 6 (token-efficient)
    # Each batch ~2k tokens → 35 reqs = ~6 calls × 2k = ~12k tokens per matrix
    # ═══════════════════════════════════════════════════════════════════════════
    all_mappings = {}
    sc_lookup = "\n".join(f"  {s['number']}. {s['name']}" for s in scenarios)
    BATCH = 6

    for batch_start in range(0, len(requirements), BATCH):
        batch = requirements[batch_start:batch_start + BATCH]

        req_lines = []
        for r in batch:
            acs_short = r["acs"][:400].replace("\n", " ")
            req_lines.append(f"[{r['req_id']}] {r['req_desc'][:80]} | ACs: {acs_short}")

        resp = await llm_call(
            "QA matrix analyst. Output ONLY JSON. Start with {.",
            (
                "Map requirements to scenarios. Return compact JSON.\n\n"
                f"SCENARIOS:\n{sc_lookup}\n\n"
                f"REQUIREMENTS:\n" + "\n".join(req_lines) + "\n\n"
                "RULES:\n"
                "- Match each req to the scenario whose name best fits its AC content.\n"
                "- Same flow → put ALL codes in ONE scenario: {\"2\":\"AC01,AC02,AC03\"}\n"
                "- Different flows → split: {\"4\":\"AC01\",\"7\":\"AC02\"}\n"
                "- Values = AC codes (AC01, AC1,AC2) NEVER scenario numbers.\n\n"
                "{\"mappings\":[{\"req_id\":\"ID\",\"scenario_mappings\":{\"sc_num\":\"AC_codes\"}}]}\n"
                "Start with {"
            ),
            2000
        )

        for m in resp.get("mappings", []):
            rid = str(m.get("req_id", ""))
            sm  = m.get("scenario_mappings", {})
            if not rid or not sm:
                continue
            clean = {}
            for k, v in sm.items():
                v_str = str(v).strip()
                if re.match(r'^\d+$', v_str):
                    v_str = f"AC{int(v_str):02d}"
                if v_str.upper().startswith("SC"):
                    v_str = "AC01"
                try:
                    sc_num = int(str(k).strip())
                    if 1 <= sc_num <= len(scenarios):
                        clean[str(sc_num)] = v_str
                except Exception:
                    pass
            if clean:
                all_mappings[rid] = clean

    # ── Fallback: partial key match for unmatched requirements ────────────────
    for req in requirements:
        rid = req["req_id"]
        if not all_mappings.get(rid):
            for k, v in all_mappings.items():
                if rid in k or k in rid:
                    all_mappings[rid] = v
                    break


    # ═══════════════════════════════════════════════════════════════════════════
    # STEP 3 — Calculate counts, remove unused scenarios, renumber
    # ═══════════════════════════════════════════════════════════════════════════
    sc_ac_count  = {str(sc["number"]): 0 for sc in scenarios}
    sc_req_count = {str(sc["number"]): 0 for sc in scenarios}

    for rid, mapping in all_mappings.items():
        for sc_num, ac_val in mapping.items():
            if ac_val and str(ac_val).strip():
                sc_req_count[sc_num] = sc_req_count.get(sc_num, 0) + 1
                sc_ac_count[sc_num]  = sc_ac_count.get(sc_num, 0) + len(
                    [x for x in str(ac_val).split(",") if x.strip()])

    # ── Remove scenarios with 0 requirements mapped — never show empty columns ─
    active_scenarios = [sc for sc in scenarios
                        if sc_req_count.get(str(sc["number"]), 0) > 0]

    if not active_scenarios:
        # Fallback: keep all if nothing mapped at all
        active_scenarios = scenarios

    # Renumber active scenarios 1..N sequentially
    old_to_new = {}  # old_number_str → new_number
    renumbered = []
    for new_idx, sc in enumerate(active_scenarios):
        new_num = new_idx + 1
        old_to_new[str(sc["number"])] = new_num
        renumbered.append({**sc, "number": new_num})

    # Remap all_mappings to use new scenario numbers
    remapped = {}
    for rid, mapping in all_mappings.items():
        new_mapping = {}
        for old_num, ac_val in mapping.items():
            if old_num in old_to_new:
                new_mapping[str(old_to_new[old_num])] = ac_val
        if new_mapping:
            remapped[rid] = new_mapping
    all_mappings = remapped
    scenarios    = renumbered

    # Recalculate counts with new numbering
    sc_ac_count  = {str(sc["number"]): 0 for sc in scenarios}
    sc_req_count = {str(sc["number"]): 0 for sc in scenarios}
    for rid, mapping in all_mappings.items():
        for sc_num, ac_val in mapping.items():
            if ac_val and str(ac_val).strip():
                sc_req_count[sc_num] = sc_req_count.get(sc_num, 0) + 1
                sc_ac_count[sc_num]  = sc_ac_count.get(sc_num, 0) + len(
                    [x for x in str(ac_val).split(",") if x.strip()])

    # ═══════════════════════════════════════════════════════════════════════════
    # STEP 4 — Build Excel matching reference format exactly
    # ═══════════════════════════════════════════════════════════════════════════
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Requirement Matrix"
    # Force auto-recalculation so formulas work on open
    wb.calculation.calcMode = "auto"

    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    def fill(c): return PatternFill("solid", fgColor=c)
    def fnt(color="000000", bold=False, sz=9):
        return Font(color=color, bold=bold, size=sz)
    ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lwrap = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    n = len(scenarios)

    # ── ROW 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8 + n)
    c1 = ws.cell(1, 1, "Requirements Traceability Matrix - E2E SIT Scenarios")
    c1.fill = fill("1F4E79"); c1.font = Font(color="FFFFFF", bold=True, size=13)
    c1.alignment = ctr

    # ── ROW 2: Pattern ────────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(2, 7, "E2E SIT Scenario").font = fnt(bold=True)
    ws.cell(2, 8, "Pattern").font          = fnt(bold=True)
    for i, sc in enumerate(scenarios):
        c = ws.cell(2, 9 + i, pattern)
        c.fill = fill("D9E1F2"); c.font = fnt(bold=True); c.alignment = ctr; c.border = bdr

    # ── ROW 3: Scenario numbers ───────────────────────────────────────────────
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    ws.cell(3, 8, "Number").font = fnt(bold=True)
    for i, sc in enumerate(scenarios):
        c = ws.cell(3, 9 + i, sc["number"])
        c.fill = fill("D9E1F2"); c.font = fnt(bold=True); c.alignment = ctr; c.border = bdr

    # ── ROW 4: Scenario names ─────────────────────────────────────────────────
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
    ws.cell(4, 8, "Name").font = fnt(bold=True)
    for i, sc in enumerate(scenarios):
        c = ws.cell(4, 9 + i, sc["name"])
        c.fill = fill("2E75B6"); c.font = fnt("FFFFFF", True)
        c.alignment = ctr; c.border = bdr

    # ── ROW 5: Section ────────────────────────────────────────────────────────
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=7)
    ws.cell(5, 1, "Requirement Details").font = fnt(bold=True)
    ws.cell(5, 8, "Section").font             = fnt(bold=True)
    for i, sc in enumerate(scenarios):
        c = ws.cell(5, 9 + i, sc.get("section", "Core Functionality"))
        c.fill = fill("4472C4"); c.font = fnt("FFFFFF", True)
        c.alignment = ctr; c.border = bdr

    # ── ROW 6: Rank — total AC count per scenario (STATIC — never shows 0) ───
    ws.cell(6, 8, "Rank").font = fnt(bold=True)
    max_ac = max(sc_ac_count.values()) if sc_ac_count else 1
    for i, sc in enumerate(scenarios):
        cnt = sc_ac_count.get(str(sc["number"]), 0)
        c = ws.cell(6, 9 + i, cnt)
        c.fill = fill("EEF2FF"); c.font = fnt(); c.alignment = ctr; c.border = bdr

    # ── ROW 7: Proposed Priority ──────────────────────────────────────────────
    ws.cell(7, 8, "Proposed Priority").font = fnt(bold=True)
    for i, sc in enumerate(scenarios):
        cnt   = sc_ac_count.get(str(sc["number"]), 0)
        ratio = cnt / (max_ac or 1)
        pri   = "High" if ratio >= 0.7 else ("Medium" if ratio >= 0.2 else "Low")
        col   = "C00000" if pri == "High" else ("ED7D31" if pri == "Medium" else "70AD47")
        c = ws.cell(7, 9 + i, pri)
        c.fill = fill("EEF2FF"); c.font = fnt(col, True); c.alignment = ctr; c.border = bdr

    # ── ROW 8: Agreed Priority (blank for manual input) ───────────────────────
    ws.cell(8, 8, "Agreed Priority").font = fnt(bold=True)
    for i in range(n):
        ws.cell(8, 9 + i, "").border = bdr

    # ── ROW 9: Stories Covered — STATIC count (no formula = no 0 bug) ────────
    ws.cell(9, 8, "Stories Covered").font = fnt(bold=True)
    for i, sc in enumerate(scenarios):
        cnt = sc_req_count.get(str(sc["number"]), 0)
        c = ws.cell(9, 9 + i, cnt)
        c.fill = fill("D9E1F2"); c.font = fnt(bold=True); c.alignment = ctr; c.border = bdr

    # ── ROW 10: Column headers + total AC count ───────────────────────────────
    for c, lbl in enumerate(["Functional Area", "BA", "Product / Sub-Area", "Req. ID",
                               "Req. Description", "Requirement", "AC/s", "ACs Covered"], 1):
        cell = ws.cell(10, c, lbl)
        cell.fill = fill("1F4E79"); cell.font = fnt("FFFFFF", True)
        cell.alignment = ctr; cell.border = bdr
    for i, sc in enumerate(scenarios):
        cnt = sc_ac_count.get(str(sc["number"]), 0)
        c = ws.cell(10, 9 + i, cnt if cnt else "")
        c.fill = fill("1F4E79"); c.font = fnt("FFFFFF", True)
        c.alignment = ctr; c.border = bdr

    # ── ROWS 11+: Data ────────────────────────────────────────────────────────
    grn = fill("92D050"); wht = fill("FFFFFF"); alt = fill("F2F2F2")
    for ri, req in enumerate(requirements):
        dr = 11 + ri
        rf = alt if ri % 2 == 0 else wht
        for c, val in enumerate([req["col_a"], req["col_b"], req["col_c"], req["req_id"],
                                   req["req_desc"], req["req_stmt"], req["acs"], req["col_h"]], 1):
            cell = ws.cell(dr, c, val)
            cell.fill = rf; cell.font = fnt(); cell.alignment = lwrap; cell.border = bdr

        mapping  = all_mappings.get(req["req_id"], {})
        total_ac = 0
        for i, sc in enumerate(scenarios):
            ac_val = str(mapping.get(str(sc["number"]), "")).strip()
            cell   = ws.cell(dr, 9 + i, ac_val if ac_val else "")
            if ac_val:
                cell.fill = grn; cell.font = fnt("000000", True)
                total_ac += len([x for x in ac_val.split(",") if x.strip()])
            else:
                cell.fill = rf; cell.font = fnt()
            cell.alignment = ctr; cell.border = bdr

        if total_ac > 0:
            ws.cell(dr, 8, total_ac).font = fnt()

    # ── Column widths & freeze ────────────────────────────────────────────────
    for c, w in enumerate([16, 10, 16, 13, 32, 38, 45, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    sc_w = max(7, 35 // max(n, 1))
    for i in range(n):
        ws.column_dimensions[get_column_letter(9 + i)].width = sc_w
    ws.row_dimensions[4].height = 55
    ws.freeze_panes = "I11"

    # ── Save & store ──────────────────────────────────────────────────────────
    out = io.BytesIO(); wb.save(out); out.seek(0); excel_bytes = out.read()
    fname = f"Matrix_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    ins = await matrix_col.insert_one({
        "user_email": user_email.lower(), "filename": fname,
        "original_filename": file.filename or "uploaded.xlsx",
        "requirements_count": len(requirements), "scenarios_count": n,
        "scenario_names": [s["name"] for s in scenarios],
        "excel_bytes": base64.b64encode(excel_bytes).decode(),
        "created_at": datetime.utcnow().isoformat(),
    })
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}",
                 "X-History-Id": str(ins.inserted_id)},
    )



@app.get("/api/llm/status")
async def llm_status():
    if groq_client is None: return {"status":"offline","reason":"GROQ_API_KEY not set"}
    try:
        await groq_client.chat.completions.create(model=GROQ_MODEL,messages=[{"role":"user","content":"hi"}],max_tokens=5)
        return {"status":"online","model":GROQ_MODEL,"provider":"Groq"}
    except Exception as e: return {"status":"offline","reason":str(e)}

@app.get("/api/health")
def health(): return {"status":"ok","model":GROQ_MODEL,"version":"v7"}
